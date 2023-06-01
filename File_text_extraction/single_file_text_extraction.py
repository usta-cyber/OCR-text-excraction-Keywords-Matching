## Text File Key Word Extract and Match
"""
Can extract and Match word within different file formats like
PDF,DOCX,XML,CSV,XSLX,JSON
"""

import os
from docx import Document
import PyPDF2
import openpyxl
import csv
import json
import xml.etree.ElementTree as ET
import pandas as pd
def read_pdf(file_path):
    with open(file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        text = ""
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()

    return text

def match_text_in_pdf(file_name,file_path, word_list):
    pdf_text = read_pdf(file_path)
    matched_words=[]
    # Convert the PDF text to lowercase for case-insensitive matching
    pdf_text_lower = pdf_text.lower()
    for word in word_list:
        # Convert the word to lowercase for case-insensitive matching
        word_lower = word.lower()

        if word_lower in pdf_text_lower:
            print(f"Match found in file: {word}")
            print(f"File Name={file_name}, path={file_path}")
            matched_words.append((word,file_name,file_path))
        # else :
        #     print("match not found")
    return matched_words

def match_text_in_word(file_name,file_path, word_list):
    matched_words=[]
    document = Document(file_path)
    for paragraph in document.paragraphs:
        text = paragraph.text
        for word in word_list:
            if word in text:
                print(f"Match found in file: {word}")
                print(f"File Name={file_name}, path={file_path}")
                matched_words.append((word,file_name,file_path))
        # else :
        #     print("match not found")
    return matched_words



def match_text_in_excel(file_name,file_path, word_list):
    matched_words=[]
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value)
                for word in word_list:
                    if word in text:
                        print(f"Match found in Excel: {word}")
                        print(f"File Name={file_name}, path={file_path}")
                        matched_words.append((word,file_name,file_path))
                    # else :
                    #     print("match not found")
    return matched_words



def match_text_in_csv(file_name,file_path, word_list):
    matched_words=[]
    with open(file_path, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            for cell in row:
                for word in word_list:
                    if word in cell:
                        print(f"File Name={file_name}, path={file_path}")
                        matched_words.append((word,file_name,file_path))
                    # else :
                    #     print("match not found")
    return matched_words




def match_text_in_json(file_name,file_path, word_list):
    matched_words=[]
    with open(file_path, 'r') as file:
        data = json.load(file)
        for key, value in data.items():
            for word in word_list:
                if word in str(value):
                    print(f"Match found in Excel: {word}")
                    print(f"File Name={file_name}, path={file_path}")
                    matched_words.append((word,file_name,file_path))
                # else :
                #     print("match not found")
    return matched_words


def match_text_in_xml(file_name,file_path, word_list):
    matched_words=[]
    tree = ET.parse(file_path)
    root = tree.getroot()
    for element in root.iter():
        for word in word_list:
            if word in element.text:
                print(f"File Name={file_name}, path={file_path}")
                matched_words.append((word,file_name,file_path))
            # else :
            #     print("match not found")
    return matched_words



def load_words_file_to_list(word_list_file_path):
    with open(word_list_file_path, 'r') as word_list_file:
        word_list = [word.strip() for word in word_list_file.readlines() if word.strip()]
    return word_list

def save_matched_words_to_excel(matched_words, output_file_path):
    df = pd.DataFrame(matched_words, columns=['Matched Word', 'File Name', 'File Path'])
    df.to_excel(output_file_path, index=False)



matched_words=[]   
# List of words to match
word_file="C:\\Users\\Tajummal\\Documents\\word_file.txt"
# Directory containing the files to process
directory = "C:\\Users\\Tajummal\\Documents\\test_doc"
# Path to the output Excel file to save matched results
output_file_path = 'C:\\Users\\Tajummal\\Documents\\matced_word_results.xlsx'

word_list=load_words_file_to_list(word_file)


for filename in os.listdir(directory):
    file_path = os.path.join(directory, filename)
    if filename.endswith(".pdf"):
       matched_words.extend(match_text_in_pdf(filename,file_path, word_list))
    elif filename.endswith(".docx"):
         matched_words.extend(match_text_in_word(filename,file_path, word_list))
    elif filename.endswith(".xlsx"):
         matched_words.extend(match_text_in_excel(filename,file_path, word_list))
    elif filename.endswith(".csv"):
         matched_words.extend(match_text_in_csv(filename,file_path, word_list))
    elif filename.endswith(".json"):
         matched_words.extend(match_text_in_json(filename,file_path, word_list))
    elif filename.endswith(".xml"):
         matched_words.extend(match_text_in_xml(filename,file_path, word_list))

# Save the matched words, file names, and file paths to an Excel file
save_matched_words_to_excel(matched_words, output_file_path)