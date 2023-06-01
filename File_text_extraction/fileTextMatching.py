import os
from docx import Document
import PyPDF2
import openpyxl
import csv
import json
import xml.etree.ElementTree as ET
import pandas as pd
def read_pdf(file_path):
    with open(file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        text = ""
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()

    return text

def match_text_in_pdf(file_name,file_path, word_list):
    pdf_text = read_pdf(file_path)
    matched_words=[]
    # Convert the PDF text to lowercase for case-insensitive matching
    pdf_text_lower = pdf_text.lower()
    for word in word_list:
        # Convert the word to lowercase for case-insensitive matching
        word_lower = word.lower()

        if word_lower in pdf_text_lower:
            print(f"Match found in file: {word}")
            print(f"File Name={file_name}, path={file_path}")
            matched_words.append((word,file_name,file_path))
        # else :
        #     print("match not found")
    return matched_words

def match_text_in_word(file_name,file_path, word_list):
    matched_words=[]
    document = Document(file_path)
    for paragraph in document.paragraphs:
        text = paragraph.text
        for word in word_list:
            if word in text:
                print(f"Match found in file: {word}")
                print(f"File Name={file_name}, path={file_path}")
                matched_words.append((word,file_name,file_path))
        # else :
        #     print("match not found")
    return matched_words



def match_text_in_excel(file_name,file_path, word_list):
    matched_words=[]
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value)
                for word in word_list:
                    if word in text:
                        print(f"Match found in Excel: {word}")
                        print(f"File Name={file_name}, path={file_path}")
                        matched_words.append((word,file_name,file_path))
                    # else :
                    #     print("match not found")
    return matched_words



def match_text_in_csv(file_name,file_path, word_list):
    matched_words=[]
    with open(file_path, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            for cell in row:
                for word in word_list:
                    if word in cell:
                        print(f"File Name={file_name}, path={file_path}")
                        matched_words.append((word,file_name,file_path))
                    # else :
                    #     print("match not found")
    return matched_words




def match_text_in_json(file_name,file_path, word_list):
    matched_words=[]
    with open(file_path, 'r') as file:
        data = json.load(file)
        for key, value in data.items():
            for word in word_list:
                if word in str(value):
                    print(f"Match found in Excel: {word}")
                    print(f"File Name={file_name}, path={file_path}")
                    matched_words.append((word,file_name,file_path))
                # else :
                #     print("match not found")
    return matched_words

def match_text_in_xml(file_name,file_path, word_list):
    matched_words=[]
    tree = ET.parse(file_path)
    root = tree.getroot()
    for element in root.iter():
        for word in word_list:
            if word in element.text:
                print(f"File Name={file_name}, path={file_path}")
                matched_words.append((word,file_name,file_path))
            # else :
            #     print("match not found")
    return matched_words
